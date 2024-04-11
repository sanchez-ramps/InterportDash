import openpyxl
import base64


from abc import ABC, abstractmethod


class Report(ABC):

    @abstractmethod
    def generate_report(self) -> openpyxl.Workbook:
        pass


class OnHandReport(Report):
    records = []

    def set_records(self, records):
        self.records = records

    def generate_report(self) -> openpyxl.Workbook:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        column_mapping_list = [
            {"label": "Status", "key": "status"},
            {"label": "Warehouse receipt #", "key": "number"},
            {"label": "Created On", "key": "created_date"},
            {"label": "Shipper", "key": "shipper"},
            {"label": "Supplier", "key": "supplier"},
            {"label": "Carrier", "key": "carrier"},
            {"label": "PRO Number", "key": ""},
            {"label": "BOL # / Supplier Reference", "key": "tracking_number"},
            {"label": "Customer Purchase Order #", "key": "purchase_order"},
            {"label": "Invoice Number", "key": "invoice"},
            {"label": "Pieces (pllts)", "key": "pieces"},
            {"label": "Weight (LB)", "key": ""},
            {"label": "Volume (FT3)", "key": ""},
            {"label": "Temperature at Receipt ©", "key": ""},
            {"label": "Activity Status", "key": ""},
            {"label": "INSTRUCTIONS", "key": "note"}
        ]

        # Set the column headings
        # Add the column headings to the first row
        for col_num, heading in enumerate(column_mapping_list, 1):
            sheet.cell(row=1, column=col_num, value=heading['label'])
            heading['column'] = col_num

        # Iterate through the data dictionaries and populate the Excel sheet
        for row_num, data_dict in enumerate(self.records, 2):
            for heading in column_mapping_list:
                key = heading.get('key', "")
                if (key):
                    column = heading.get('column', "")
                    # Get the value from the dictionary or use an empty string if the key is missing
                    value = data_dict.get(key, "")
                    sheet.cell(row=row_num, column=column, value=value)


        print(workbook.save("./data.xlsx"))
        workbook.close()
        return workbook


def on_hand(records):
    # Create a new Excel workbook

    # Select the active sheet (the first sheet by default)

    column_headings = [
        "Status", "Warehouse receipt #", "Created On", "Shipper", "Supplier",
        "Carrier", "PRO Number", "BOL # / Supplier Reference", "Customer Purchase Order #",
        "Invoice Number", "Pieces (pllts)", "Weight (LB)", "Volume (FT3)",
        "Temperature at Receipt ©", "Activity Status", "INSTRUCTIONS"
    ]

    # Add the column headings to the first row
    for col_num, heading in enumerate(column_headings, 1):
        sheet.cell(row=1, column=col_num, value=heading)
